import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

class ExcelExporter:
    def __init__(self):
        # Стили для Excel
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.cell_font = Font(name='Arial', size=10)
        self.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
    def create_excel_download(self, dataframe):
        """Создает Excel файл для скачивания в Streamlit"""
        output = BytesIO()
        
        # Создаем workbook
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Записываем основные данные
            dataframe.to_excel(writer, index=False, sheet_name='Данные')
            
            # Получаем worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Данные']
            
            # Применяем стили
            self._apply_styles(worksheet, len(dataframe))
            
            # Автоматическая ширина колонок
            self._auto_adjust_column_width(worksheet)
            
            # Добавляем лист со статистикой
            self._add_statistics_sheet(workbook, dataframe)
        
        output.seek(0)
        return output.getvalue()
    
    def export_to_excel(self, dataframe, filename="extracted_data.xlsx"):
        """Экспортирует DataFrame в Excel файл (для локального сохранения)"""
        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Извлеченные данные"
        
        # Добавляем данные
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws.append(r)
        
        # Применяем стили
        self._apply_styles(ws, len(dataframe))
        
        # Автоматическая ширина колонок
        self._auto_adjust_column_width(ws)
        
        # Добавляем лист со статистикой
        self._add_statistics_sheet(wb, dataframe)
        
        # Сохраняем файл
        wb.save(filename)
        wb.close()
        
        return filename
    
    def _apply_styles(self, worksheet, row_count):
        """Применяет стили к таблице"""
        # Стили для заголовков
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Стили для данных
        for row_num in range(2, row_count + 2):
            for col_num in range(1, 4):  # 3 колонки
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = self.cell_font
                cell.border = self.border
                
                # Выравнивание
                if col_num == 1:  # Артикул
                    cell.alignment = Alignment(horizontal='center')
                elif col_num == 2:  # Наименование
                    cell.alignment = Alignment(horizontal='left')
                else:  # Количество
                    cell.alignment = Alignment(horizontal='center')
    
    def _auto_adjust_column_width(self, worksheet):
        """Автоматически настраивает ширину колонок"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_statistics_sheet(self, workbook, dataframe):
        """Добавляет лист со статистикой"""
        stats_sheet = workbook.create_sheet(title="Статистика")
        
        # Общая статистика
        stats = [
            ["Показатель", "Значение"],
            ["Всего позиций", len(dataframe)],
            ["Уникальных артикулов", dataframe['Артикул'].nunique()],
            ["Позиций с артикулами", len(dataframe[dataframe['Артикул'] != ''])],
            ["Позиций с наименованиями", len(dataframe[dataframe['Наименование'] != ''])],
            ["Позиций с количеством", len(dataframe[dataframe['Количество'] != ''])],
            ["", ""],
            ["Топ артикулы", ""],
        ]
        
        # Добавляем топ артикулы
        top_articles = dataframe[dataframe['Артикул'] != '']['Артикул'].value_counts().head(5)
        for article, count in top_articles.items():
            stats.append([f"  {article}", f"{count} раз"])
        
        # Записываем статистику
        for row in stats:
            stats_sheet.append(row)
        
        # Применяем стили к статистике
        for row_num in range(1, len(stats) + 1):
            for col_num in range(1, 3):
                cell = stats_sheet.cell(row=row_num, column=col_num)
                cell.border = self.border
                
                if row_num == 1:  # Заголовок
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center')
                elif row_num == 8:  # Подзаголовок "Топ артикулы"
                    cell.font = Font(name='Arial', size=11, bold=True)
                else:
                    cell.font = self.cell_font
        
        # Автоматическая ширина колонок для статистики
        self._auto_adjust_column_width(stats_sheet)
    
    def export_sample_data(self):
        """Создает пример данных для тестирования"""
        sample_data = [
            {"Артикул": "A001-2024", "Наименование": "Болт крепёжный М8х40", "Количество": "50 шт"},
            {"Артикул": "B152-ST", "Наименование": "Гайка шестигранная М10", "Количество": "100 шт"},
            {"Артикул": "C003-AL", "Наименование": "Шайба алюминиевая Ø12", "Количество": "25 шт"},
            {"Артикул": "D-445", "Наименование": "Винт самонарезающий 4x16", "Количество": "200 шт"},
            {"Артикул": "E12-PRO", "Наименование": "Дюбель пластиковый 6x30", "Количество": "150 шт"},
        ]
        
        df = pd.DataFrame(sample_data)
        return self.create_excel_download(df)